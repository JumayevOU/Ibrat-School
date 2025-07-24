import os
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

FOLDER = "xisobot"

def save_user_data(fullname, phone):

    if not os.path.exists(FOLDER):
        os.makedirs(FOLDER)


    today_str = datetime.today().strftime("%Y.%m.%d")
    file_path = os.path.join(FOLDER, f"{today_str}.xlsx")

    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Anketalar"
        sheet.append(["Ism Familiya", "Telefon raqam"])
        wb.save(file_path)


    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    sheet.append([fullname, phone])

    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 2  

    wb.save(file_path)
